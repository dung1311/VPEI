from datetime import datetime
from io import BytesIO
import os
from sqlalchemy.orm import Session
from sqlalchemy import desc
from fastapi import HTTPException, status
from typing import Optional, List, Dict, Any
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from models.electrical_item import ElectricalItem, ItemLocation
from services.scope2_activity_service import record_activity

EF = 0.6235


def _format_activity_date(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return raw

def _parse_entry_date(value: str) -> datetime:
    value = (value or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    raise ValueError("entry_date must be yyyy-mm-dd or dd/mm/yyyy")


def _safe_pdf_text(value: Any) -> str:
    return str(value or "")


def _get_pdf_fonts() -> tuple[str, str]:
    regular_candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    ]
    bold_candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    ]

    regular_path = next((p for p in regular_candidates if os.path.exists(p)), None)
    bold_path = next((p for p in bold_candidates if os.path.exists(p)), None)

    if regular_path and "DejaVuSans" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans", regular_path))
    if bold_path and "DejaVuSans-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", bold_path))

    if (
        "DejaVuSans" in pdfmetrics.getRegisteredFontNames()
        and "DejaVuSans-Bold" in pdfmetrics.getRegisteredFontNames()
    ):
        return ("DejaVuSans", "DejaVuSans-Bold")

    return ("Helvetica", "Helvetica-Bold")


def _normalize_location(value: str) -> ItemLocation:
    raw = (value or "").strip()
    for loc in ItemLocation:
        if raw == loc.value:
            return loc
    return ItemLocation.MAIN_PORT


def _parse_item_date(item: ElectricalItem) -> Optional[datetime]:
    try:
        return _parse_entry_date(item.period_value or "")
    except ValueError:
        return None


def _filter_items_by_period(
    items: List[ElectricalItem],
    mode: Optional[str] = None,
    bucket: Optional[str] = None,
    until_now: bool = False,
) -> List[ElectricalItem]:
    mode = (mode or "").strip().lower()
    bucket = (bucket or "").strip()
    now = datetime.now()

    def matches_mode(dt: datetime) -> bool:
        if not mode or not bucket:
            return True
        if mode == "day":
            return dt.strftime("%Y-%m-%d") == bucket
        if mode == "month":
            return dt.strftime("%Y-%m") == bucket
        if mode == "quarter":
            quarter = ((dt.month - 1) // 3) + 1
            return f"{dt.year}-Q{quarter}" == bucket
        if mode == "year":
            return str(dt.year) == bucket
        return True

    out: List[ElectricalItem] = []
    for item in items:
        dt = _parse_item_date(item)
        if dt is None:
            continue
        if until_now and dt > now:
            continue
        if not matches_mode(dt):
            continue
        out.append(item)
    return out

def get_scope2_categories(db: Session) -> List[Dict[str, Any]]:
    items_db = db.query(ElectricalItem).order_by(desc(ElectricalItem.id)).all()
    categories = []
    
    for item in items_db:
        entry_date_raw = (item.period_value or "").strip()
        try:
            dt = _parse_entry_date(entry_date_raw)
            entry_date = dt.strftime("%Y-%m-%d")
        except ValueError:
            entry_date = datetime.now().strftime("%Y-%m-%d")

        categories.append({
            "id": item.id,
            "name": item.name,
            "capacity": item.power,
            "area": item.location.value if item.location else "Cảng chính",
            "entry_date": entry_date,
            "kwh": item.power,
            "note": item.description or ""
        })
    return categories

def create_electrical_item(item_data, db: Session, actor: str = "system") -> Dict[str, Any]:
    try:
        dt = _parse_entry_date(item_data.entry_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="entry_date must be yyyy-mm-dd or dd/mm/yyyy",
        )

    try:
        location_enum = ItemLocation(item_data.location)
    except ValueError:
        location_enum = ItemLocation.MAIN_PORT

    period_val = dt.strftime("%Y-%m-%d")
    
    month_prefix = dt.strftime("%Y-%m-")
    dup_item = db.query(ElectricalItem).filter(
        ElectricalItem.name == item_data.name,
        ElectricalItem.period_value.like(f"{month_prefix}%")
    ).first()

    if dup_item:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Đã có thiết bị này trong tháng, vui lòng cập nhật bằng tay."
        )

    db_item = ElectricalItem(
        name=item_data.name,
        power=item_data.power,
        location=location_enum,
        description=item_data.description,
        period_type="day",
        period_value=period_val
    )
    db.add(db_item)
    action_desc = "Thêm item (nhập thủ công)"
        
    db.commit()
    db.refresh(db_item)

    record_activity(
        db,
        actor,
        action_desc,
        f"{db_item.name} - {db_item.power:g} kWh - {_format_activity_date(db_item.period_value)}",
    )
    
    return {
        "id": db_item.id,
        "name": db_item.name,
        "capacity": db_item.power,
        "area": db_item.location.value if db_item.location else "Cảng chính",
        "entry_date": db_item.period_value,
        "kwh": db_item.power,
        "note": db_item.description or ""
    }

def update_electrical_item(item_id: int, item_data, db: Session, actor: str = "system") -> Dict[str, Any]:
    db_item = db.query(ElectricalItem).filter(ElectricalItem.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

    try:
        dt = _parse_entry_date(item_data.entry_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="entry_date must be yyyy-mm-dd or dd/mm/yyyy",
        )

    try:
        location_enum = ItemLocation(item_data.location)
    except ValueError:
        location_enum = ItemLocation.MAIN_PORT

    reason = (item_data.update_reason or "").strip()
    if not reason:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="update_reason is required",
        )

    db_item.name = item_data.name
    db_item.power = item_data.power
    db_item.location = location_enum
    db_item.description = item_data.description
    db_item.period_type = "day"
    db_item.period_value = dt.strftime("%Y-%m-%d")

    db.commit()
    db.refresh(db_item)

    record_activity(
        db,
        actor,
        "Sửa item",
        f"{db_item.name} - {db_item.power:g} kWh - {_format_activity_date(db_item.period_value)} - Lý do: {reason}",
    )

    return {
        "id": db_item.id,
        "name": db_item.name,
        "capacity": db_item.power,
        "area": db_item.location.value if db_item.location else "Cảng chính",
        "entry_date": db_item.period_value,
        "kwh": db_item.power,
        "note": db_item.description or "",
    }

def delete_electrical_item(item_id: int, db: Session, actor: str = "system") -> Dict[str, Any]:
    db_item = db.query(ElectricalItem).filter(ElectricalItem.id == item_id).first()
    if not db_item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Item not found",
        )

    deleted_snapshot = {
        "name": db_item.name,
        "power": db_item.power,
        "period_value": db_item.period_value or "",
    }
    db.delete(db_item)
    db.commit()

    record_activity(
        db,
        actor,
        "Xóa item",
        f"{deleted_snapshot['name']} - {deleted_snapshot['power']:g} kWh - {_format_activity_date(deleted_snapshot['period_value'])}",
    )

    return {"ok": True, "deleted_id": item_id}


def import_scope2_items_from_excel(file_bytes: bytes, db: Session, actor: str = "system") -> Dict[str, Any]:
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": True, "imported": 0, "failed": 0, "errors": []}

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    idx_by_header = {h: i for i, h in enumerate(headers) if h}

    def find_index(*candidates: str) -> int:
        for c in candidates:
            if c in idx_by_header:
                return idx_by_header[c]
        return -1

    idx_name = find_index("name", "ten hang muc", "ten", "hạng mục", "tên hạng mục")
    idx_power = find_index("power", "cong suat", "công suất", "capacity")
    idx_location = find_index("location", "khu vuc", "khu vực", "area")
    idx_entry_date = find_index("entry_date", "ngay nhap", "ngày nhập", "date")
    idx_note = find_index("description", "note", "ghi chu", "ghi chú")

    if min(idx_name, idx_power, idx_location, idx_entry_date) < 0:
        raise HTTPException(
            status_code=400,
            detail="Excel must include columns: name, power, location, entry_date",
        )

    imported = 0
    failed = 0
    errors: List[str] = []

    for row_no, row in enumerate(rows[1:], start=2):
        try:
            name = str(row[idx_name] or "").strip()
            power = float(row[idx_power])
            location = _normalize_location(str(row[idx_location] or ""))
            entry_raw = str(row[idx_entry_date] or "").strip()
            note = str(row[idx_note] or "").strip() if idx_note >= 0 else ""

            if not name:
                raise ValueError("name is required")
            if power <= 0:
                raise ValueError("power must be > 0")

            try:
                dt = _parse_entry_date(entry_raw)
            except ValueError as ex:
                raise ValueError(str(ex))

            period_val = dt.strftime("%Y-%m-%d")
            existing_item = db.query(ElectricalItem).filter(
                ElectricalItem.name == name,
                ElectricalItem.location == location,
                ElectricalItem.period_value == period_val
            ).first()

            if existing_item:
                existing_item.power += power
                if note:
                    existing_item.description = note
                imported += 1
            else:
                db_item = ElectricalItem(
                    name=name,
                    power=power,
                    location=location,
                    description=note,
                    period_type="day",
                    period_value=period_val,
                )
                db.add(db_item)
                imported += 1
        except Exception as ex:
            failed += 1
            errors.append(f"Row {row_no}: {ex}")

    db.commit()
    record_activity(
        db,
        actor,
        "Thêm item (nhập Excel)",
        f"Imported {imported} item(s), failed {failed} item(s)",
    )
    return {"ok": True, "imported": imported, "failed": failed, "errors": errors}


def export_scope2_items_excel(
    db: Session,
    mode: Optional[str] = None,
    bucket: Optional[str] = None,
    until_now: bool = False,
    actor: str = "system",
) -> bytes:
    items_db = db.query(ElectricalItem).order_by(desc(ElectricalItem.id)).all()
    items_db = _filter_items_by_period(items_db, mode=mode, bucket=bucket, until_now=until_now)

    wb = Workbook()
    ws = wb.active
    ws.title = "Scope2"
    ws.append(["STT", "Name", "Điện năng(kWh)", "Location", "Entry Date", "Description", "kWh"])
    count = 0
    for item in items_db:
        count += 1
        ws.append([
            str(count),
            item.name,
            item.power,
            item.location.value if item.location else "Cảng chính",
            item.period_value or "",
            item.description or "",
            item.power,
        ])

    out = BytesIO()
    wb.save(out)
    record_activity(
        db,
        actor,
        "Xuất report Excel",
        f"{len(items_db)} item(s) - mode={mode or 'all'} - bucket={bucket or 'all'} - until_now={until_now}",
    )
    return out.getvalue()


def export_scope2_import_template_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scope2_Template"
    ws.append(["name", "power", "location", "entry_date", "description"])
    ws.append(["Hệ thống điều hòa trung tâm", 350, "Văn phòng", "2026-03-27", "Ví dụ dữ liệu mẫu"])
    ws.append(["Chiếu sáng khu bến", 120, "Cảng chính", "2026-03-27", ""])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def export_scope2_items_pdf(
    db: Session,
    mode: Optional[str] = None,
    bucket: Optional[str] = None,
    until_now: bool = False,
    actor: str = "system",
) -> bytes:
    items_db = db.query(ElectricalItem).order_by(desc(ElectricalItem.id)).all()
    items_db = _filter_items_by_period(items_db, mode=mode, bucket=bucket, until_now=until_now)
    normal_font, bold_font = _get_pdf_fonts()

    out = BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = bold_font
    styles["Normal"].fontName = normal_font

    elems = [
        Paragraph(_safe_pdf_text("Báo cáo Scope 2 - Danh mục điện năng"), styles["Title"]),
        Spacer(1, 10),
        # Paragraph(_safe_pdf_text(f"Generated at: {datetime.now().strftime('%d/%m/%Y %H:%M')}"), styles["Normal"]),
        Spacer(1, 12),
    ]

    data = [["STT", "Name", "Điện năng(kWh)", "Location", "Entry Date", "kWh"]]
    count = 0
    for item in items_db:
        count += 1
        data.append([
            str(count),
            _safe_pdf_text(item.name),
            f"{item.power:g}",
            _safe_pdf_text(item.location.value if item.location else "Cảng chính"),
            _safe_pdf_text(item.period_value or ""),
            f"{item.power:g}",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d1f3c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTNAME", (0, 1), (-1, -1), normal_font),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.Color(0.97, 0.98, 1)]),
    ]))

    elems.append(table)
    doc.build(elems)
    record_activity(
        db,
        actor,
        "Xuất report PDF",
        f"{len(items_db)} item(s) - mode={mode or 'all'} - bucket={bucket or 'all'} - until_now={until_now}",
    )
    return out.getvalue()
